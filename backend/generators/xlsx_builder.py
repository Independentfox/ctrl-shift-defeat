import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from models.context_object import StartupContextObject


HEADER_FILL = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def build_financial_model(data: dict, context: StartupContextObject) -> bytes:
    wb = Workbook()

    # Sheet 1: Revenue Projections
    ws = wb.active
    ws.title = "Revenue Projections"
    _add_header(ws, ["Period", "Revenue", "Description"], row=1)

    rev = data.get("revenue_projections", {})
    row = 2
    for period_key in ["month_1_6", "month_7_12", "month_13_18"]:
        period = rev.get(period_key, {})
        label = period_key.replace("_", " ").replace("month", "Month").title()
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=period.get("amount", "N/A"))
        ws.cell(row=row, column=3, value=period.get("description", ""))
        row += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 50

    # Sheet 2: Cost Structure
    ws2 = wb.create_sheet("Cost Structure")
    _add_header(ws2, ["Category", "Monthly Cost"], row=1)

    costs = data.get("cost_structure", {})
    row = 2
    for category, amount in costs.items():
        ws2.cell(row=row, column=1, value=category.replace("_", " ").title())
        ws2.cell(row=row, column=2, value=str(amount))
        row += 1

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 25

    # Sheet 3: Funding Analysis
    ws3 = wb.create_sheet("Funding Analysis")
    _add_header(ws3, ["Metric", "Value"], row=1)

    funding = data.get("funding_analysis", {})
    row = 2
    for key, value in funding.items():
        ws3.cell(row=row, column=1, value=key.replace("_", " ").title())
        ws3.cell(row=row, column=2, value=str(value))
        row += 1

    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 25

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _add_header(ws, headers: list[str], row: int = 1):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
